"""
Module Utilitaires - Tanmia Scraper MVP v1.8
Export Excel enrichi avec infos parsing fichiers

CHANGELOG v1.8:
- NOUVEAU: Colonne Emails_Fichiers pour emails extraits des PDF/DOC
- NOUVEAU: Statistiques fichiers parsés vs détectés
- NOUVEAU: Indicateur visuel dans Excel pour fichiers avec contenu
- Amélioration: Fusion automatique emails page + fichiers
"""
import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Tuple, Optional


# ============================================================================
# FORMATAGE FICHIERS ATTACHÉS (v1.8 enrichi)
# ============================================================================

def format_fichiers_attaches(fichiers: List[Dict]) -> Tuple[str, str, int, str, int]:
    """
    Formate la liste des fichiers attachés pour l'export Excel (v1.8).
    
    Args:
        fichiers: Liste de dicts {nom, url, type, contenu_texte, emails_fichier}
    
    Returns:
        Tuple (noms_str, urls_str, count, emails_fichiers_str, nb_parses)
    """
    if not fichiers or len(fichiers) == 0:
        return '', '', 0, '', 0
    
    noms = [f.get('nom', 'Sans nom') for f in fichiers]
    urls = [f.get('url', '') for f in fichiers]
    
    # Emails extraits des fichiers (v1.8)
    all_emails_fichiers = []
    nb_parses = 0
    
    for f in fichiers:
        emails_f = f.get('emails_fichier', [])
        all_emails_fichiers.extend(emails_f)
        
        if f.get('contenu_texte'):
            nb_parses += 1
    
    noms_str = ', '.join(noms)
    urls_str = '\n'.join(urls)
    emails_fichiers_str = ', '.join(set(all_emails_fichiers))
    
    return noms_str, urls_str, len(fichiers), emails_fichiers_str, nb_parses


# ============================================================================
# EXPORT EXCEL (v1.8 enrichi)
# ============================================================================

def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Exporte un DataFrame vers Excel avec formatage v1.8.
    
    Inclut colonnes emails fichiers et indicateurs parsing.
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Opportunités')
        
        worksheet = writer.sheets['Opportunités']
        
        # Auto-ajustement colonnes
        for idx, column in enumerate(df.columns):
            max_length = df[column].astype(str).map(len).max()
            header_length = len(column)
            column_length = max(max_length, header_length)
            
            if column in ['URL', 'Liens_Fichiers']:
                column_length = min(column_length + 2, 60)
            elif column in ['Résumé', 'Mots-clés', 'Fichiers', 'Emails_Fichiers']:
                column_length = min(column_length + 2, 45)
            else:
                column_length = min(column_length + 2, 30)
            
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = column_length
        
        # Formatage header
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Formatage contenu
        content_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.alignment = content_alignment
                cell.border = thin_border
        
        # Surlignage lignes avec fichiers parsés (v1.8)
        if 'Nb_Parses' in df.columns:
            highlight_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Vert clair
            highlight_fill_partial = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Jaune
            
            nb_parses_col = df.columns.get_loc('Nb_Parses') + 1
            
            for row_idx in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_idx, column=nb_parses_col).value
                try:
                    if cell_value and int(cell_value) > 0:
                        # Vert: fichiers parsés avec succès
                        for col_idx in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row_idx, column=col_idx).fill = highlight_fill
                except:
                    pass
        
        # Surlignage emails fichiers (v1.8)
        if 'Emails_Fichiers' in df.columns:
            email_highlight = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Bleu clair
            emails_col = df.columns.get_loc('Emails_Fichiers') + 1
            
            for row_idx in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_idx, column=emails_col).value
                if cell_value and len(str(cell_value).strip()) > 0:
                    worksheet.cell(row=row_idx, column=emails_col).fill = email_highlight
        
        worksheet.freeze_panes = 'A2'
        
        for row in range(2, len(df) + 2):
            worksheet.row_dimensions[row].height = 45
    
    output.seek(0)
    return output.getvalue()


def get_column_letter(col_idx: int) -> str:
    """Convertit index colonne en lettre Excel."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def create_export_filename(url_type: str) -> str:
    """Génère nom de fichier avec timestamp."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"tanmia_{url_type}_{timestamp}.xlsx"


# ============================================================================
# STATISTIQUES (v1.8 enrichi)
# ============================================================================

def calculate_statistics(df: pd.DataFrame) -> Dict:
    """
    Calcule des statistiques complètes (v1.8).
    
    Inclut stats sur fichiers parsés et emails extraits.
    """
    stats = {
        'total_opportunites': len(df),
        'avec_email': 0,
        'sans_email': 0,
        'taux_email': 0.0,
        'secteurs_uniques': 0,
        'top_secteurs': [],
        'top_organisations': [],
        # v1.7
        'avec_fichiers': 0,
        'total_fichiers': 0,
        'taux_fichiers': 0.0,
        # v1.8 NOUVEAU
        'fichiers_parses': 0,
        'emails_from_files': 0,
        'taux_parsing': 0.0
    }
    
    if len(df) == 0:
        return stats
    
    # Emails (page)
    if 'Email' in df.columns:
        avec_email = df['Email'].apply(
            lambda x: len(str(x).strip()) > 0 and str(x).strip() != ''
        ).sum()
        stats['avec_email'] = int(avec_email)
        stats['sans_email'] = len(df) - stats['avec_email']
        stats['taux_email'] = round((stats['avec_email'] / len(df)) * 100, 1)
    
    # Secteurs
    if 'Secteur' in df.columns:
        secteurs = df['Secteur'].value_counts()
        stats['secteurs_uniques'] = len(secteurs)
        stats['top_secteurs'] = secteurs.head(5).to_dict()
    
    # Organisations
    if 'Organisation' in df.columns:
        orgs = df['Organisation'].value_counts()
        stats['top_organisations'] = orgs.head(5).to_dict()
    
    # Fichiers (v1.7)
    if 'Nb_Fichiers' in df.columns:
        avec_fichiers = df['Nb_Fichiers'].apply(
            lambda x: int(x) > 0 if pd.notna(x) else False
        ).sum()
        total_fichiers = df['Nb_Fichiers'].apply(
            lambda x: int(x) if pd.notna(x) else 0
        ).sum()
        
        stats['avec_fichiers'] = int(avec_fichiers)
        stats['total_fichiers'] = int(total_fichiers)
        stats['taux_fichiers'] = round((stats['avec_fichiers'] / len(df)) * 100, 1)
    
    # Fichiers parsés (v1.8)
    if 'Nb_Parses' in df.columns:
        total_parses = df['Nb_Parses'].apply(
            lambda x: int(x) if pd.notna(x) else 0
        ).sum()
        stats['fichiers_parses'] = int(total_parses)
        
        if stats['total_fichiers'] > 0:
            stats['taux_parsing'] = round((stats['fichiers_parses'] / stats['total_fichiers']) * 100, 1)
    
    # Emails from files (v1.8)
    if 'Emails_Fichiers' in df.columns:
        avec_emails_fichiers = df['Emails_Fichiers'].apply(
            lambda x: len(str(x).strip()) > 0 and str(x).strip() != ''
        ).sum()
        stats['emails_from_files'] = int(avec_emails_fichiers)
    
    return stats


# ============================================================================
# FORMATAGE DONNÉES
# ============================================================================

def format_email_column(emails: List[str]) -> str:
    """Formate liste d'emails pour affichage."""
    if not emails or len(emails) == 0:
        return ""
    unique_emails = sorted(set(emails))
    return ", ".join(unique_emails)


def format_keywords_column(keywords: List[str]) -> str:
    """Formate liste de mots-clés pour affichage."""
    if not keywords or len(keywords) == 0:
        return ""
    keywords = keywords[:8]
    return ", ".join(keywords)


def truncate_text(text: str, max_length: int = 100) -> str:
    """Tronque un texte avec ellipse."""
    if not text:
        return ""
    text = str(text).strip()
    if len(text) <= max_length:
        return text
    return text[:max_length-3] + "..."


# ============================================================================
# FUSION RÉSULTATS (v1.8 enrichi)
# ============================================================================

def merge_analysis_results(
    scraped_data: List[Dict],
    analysis_results: List[Dict]
) -> List[Dict]:
    """
    Fusionne données scrapées et résultats IA (v1.8).
    
    Inclut emails fichiers et stats parsing.
    """
    merged = []
    
    for scraped, analysis in zip(scraped_data, analysis_results):
        # Fichiers avec infos parsing (v1.8)
        fichiers = scraped.get('fichiers_attaches', [])
        noms_fichiers, urls_fichiers, nb_fichiers, emails_fichiers, nb_parses = format_fichiers_attaches(fichiers)
        
        # Emails: fusionner page + fichiers (v1.8)
        emails_page = analysis.get('emails', [])
        emails_from_files = scraped.get('emails_from_files', [])
        all_emails = list(set(emails_page + emails_from_files))
        
        merged_item = {
            # Données brutes
            'URL': scraped.get('url', ''),
            'Titre': scraped.get('titre', ''),
            'Date': scraped.get('date', ''),
            
            # Données analysées
            'Organisation': analysis.get('organisation', scraped.get('organisation', 'Non spécifié')),
            'Email': format_email_column(all_emails),  # Fusionnés v1.8
            'Secteur': analysis.get('secteur', ''),
            'Type': analysis.get('type_opportunite', ''),
            'Localisation': analysis.get('localisation', ''),
            'Résumé': analysis.get('resume', ''),
            'Mots-clés': format_keywords_column(analysis.get('mots_cles', [])),
            
            # Fichiers (v1.7)
            'Fichiers': noms_fichiers,
            'Liens_Fichiers': urls_fichiers,
            'Nb_Fichiers': nb_fichiers,
            
            # Parsing (v1.8 NOUVEAU)
            'Emails_Fichiers': emails_fichiers,
            'Nb_Parses': nb_parses
        }
        
        merged.append(merged_item)
    
    return merged


# ============================================================================
# VALIDATION
# ============================================================================

def validate_opportunity_data(data: Dict) -> bool:
    """Valide qu'une opportunité a les champs minimum."""
    base_fields = ['url', 'titre', 'organisation', 'date']
    
    for field in base_fields:
        if field not in data or not data[field]:
            return False
    
    if 'texte_complet' not in data and 'description' not in data:
        return False
    
    return True


def clean_text(text: str) -> str:
    """Nettoie un texte."""
    if not text:
        return ""
    text = ' '.join(text.split())
    text = '\n'.join(line for line in text.split('\n') if line.strip())
    return text.strip()


# ============================================================================
# TEST
# ============================================================================

def test_utils():
    """Test utils v1.8."""
    print("🧪 TEST UTILS v1.8 (avec parsing fichiers)")
    print("=" * 60)
    
    # Test formatage fichiers avec contenu (v1.8)
    print("\n1. Formatage fichiers avec emails:")
    fichiers = [
        {
            'nom': 'TDR.pdf', 
            'url': 'https://example.com/tdr.pdf', 
            'type': 'pdf',
            'contenu_texte': 'Contenu du TDR...',
            'emails_fichier': ['contact@tdr.org', 'info@tdr.org']
        },
        {
            'nom': 'Budget.xlsx',
            'url': 'https://example.com/budget.xlsx',
            'type': 'xlsx',
            'contenu_texte': '',  # Non parsé
            'emails_fichier': []
        }
    ]
    noms, urls, count, emails_f, nb_parses = format_fichiers_attaches(fichiers)
    print(f"   Fichiers: {count}, Parsés: {nb_parses}")
    print(f"   Emails fichiers: {emails_f}")
    
    # Test merge avec fusion emails (v1.8)
    print("\n2. Merge avec fusion emails:")
    scraped = [{
        'url': 'https://test.com',
        'titre': 'Test',
        'date': '2026-02-11',
        'organisation': 'Test Org',
        'texte_complet': 'Desc...',
        'fichiers_attaches': fichiers,
        'emails_from_files': ['contact@tdr.org', 'info@tdr.org']
    }]
    analysis = [{
        'organisation': 'Test Org',
        'emails': ['page@test.org'],  # Email de la page
        'secteur': 'Éducation',
        'type_opportunite': 'Appel d\'offres',
        'localisation': 'Rabat',
        'resume': 'Résumé...',
        'mots_cles': ['M&E']
    }]
    merged = merge_analysis_results(scraped, analysis)
    print(f"   Email (fusionné): {merged[0]['Email']}")
    print(f"   Emails_Fichiers: {merged[0]['Emails_Fichiers']}")
    print(f"   Nb_Parses: {merged[0]['Nb_Parses']}")
    
    # Test stats v1.8
    print("\n3. Statistiques v1.8:")
    df = pd.DataFrame(merged)
    stats = calculate_statistics(df)
    print(f"   Fichiers parsés: {stats['fichiers_parses']}")
    print(f"   Taux parsing: {stats['taux_parsing']}%")
    print(f"   Opportunités avec emails fichiers: {stats['emails_from_files']}")
    
    # Test export
    print("\n4. Export Excel:")
    excel_bytes = export_to_excel(df)
    print(f"   Taille: {len(excel_bytes)} bytes")
    print(f"   Colonnes: {list(df.columns)}")
    
    print("\n✅ Test terminé")


if __name__ == "__main__":
    test_utils()
